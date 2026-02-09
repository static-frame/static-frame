from __future__ import annotations

import datetime
from functools import partial

import numpy as np
import typing_extensions as tp

from static_frame.core.container_util import (
    apex_to_name,
    index_from_optional_constructors,
)

# from static_frame.core.doc_str import doc_inject
from static_frame.core.frame import Frame
from static_frame.core.index import Index
from static_frame.core.index_hierarchy import IndexHierarchy
from static_frame.core.store import Store, store_coherent_non_write, store_coherent_write
from static_frame.core.store_config import StoreConfigXLSX
from static_frame.core.util import (
    BOOL_TYPES,
    COMPLEX_TYPES,
    DTYPE_BOOL,
    DTYPE_INEXACT_KINDS,
    DTYPE_INT_KINDS,
    DTYPE_OBJECT,
    DTYPE_STR_KINDS,
    NUMERIC_TYPES,
    STORE_LABEL_DEFAULT,
    TIndexCtor,
    TLabel,
    TNDArray1DBool,
    TNDArray2DBool,
    array1d_to_last_contiguous_to_edge,
)

if tp.TYPE_CHECKING:
    from openpyxl import Workbook as WorkbookOpenpyxl
    from xlsxwriter.format import Format
    from xlsxwriter.workbook import Workbook
    from xlsxwriter.worksheet import Worksheet

    TDtypeAny = np.dtype[tp.Any]

TFrameAny = Frame[tp.Any, tp.Any, tp.Unpack[tuple[tp.Any, ...]]]

MAX_XLSX_ROWS = 1048576
MAX_XLSX_COLUMNS = 16384  # 1024 on libre office


class FormatDefaults:
    @staticmethod
    def label(f: Format) -> Format:
        f.set_bold()
        return f

    @staticmethod
    def date(f: Format) -> Format:
        f.set_num_format('yyyy-mm-dd')
        return f

    @staticmethod
    def datetime(f: Format) -> Format:
        f.set_num_format('yyyy-mm-ddThh:mm:ss.000')  # ISO 8601 requires the T
        return f

    @staticmethod
    def get_format_or_default(
        workbook: Workbook,
        # format_specifier: tp.Optional[tp.Dict[str, tp.Any]],
        format_funcs: tp.Iterable[tp.Callable[[Format], Format]],
    ) -> Format:
        # if format_specifier:
        #     return workbook.add_format(format_specifier)
        f = workbook.add_format()
        for func in format_funcs:
            f = func(f)
        return f


class TWriter(tp.Protocol):
    def __call__(
        self,
        row: int,
        col: int,
        value: tp.Any,
        format_date: Format,
        format_datetime: Format,
        format_cell: Format | None = None,
    ) -> tp.Any: ...


class StoreXLSX(Store[StoreConfigXLSX]):
    _EXT: frozenset[str] = frozenset(('.xlsx',))
    _STORE_CONFIG_CLASS = StoreConfigXLSX

    @staticmethod
    def _dtype_to_writer_attr(dtype: TDtypeAny) -> tuple[str, bool]:
        """
        Return a pair of writer function, Boolean, where Boolean denotes if replacements need be applied.
        """
        kind = dtype.kind

        # NOTE: xlsxwriter cannot handle datetime64, raises TypeError('Unknown or unsupported datetime type')
        # if kind in DTYPE_NAT_KINDS and dtype != DT64_MONTH and dtype != DT64_YEAR:
        #     return 'write_datetime', True

        if dtype == DTYPE_BOOL:
            return 'write_boolean', False

        if kind in DTYPE_STR_KINDS:
            return 'write_string', False

        if kind in DTYPE_INT_KINDS:
            return 'write_number', False

        if kind in DTYPE_INEXACT_KINDS:
            return 'write_number', True

        return 'write', True

    @classmethod
    def _get_writer(cls, dtype: TDtypeAny, ws: Worksheet) -> TWriter:
        """
        Return a writer function of the passed in Worksheet.
        """
        assert isinstance(dtype, np.dtype)

        writer_attr, _ = cls._dtype_to_writer_attr(dtype)
        writer_native = getattr(ws, writer_attr)

        def writer(
            row: int,
            col: int,
            value: tp.Any,
            format_date: Format,
            format_datetime: Format,
            format_cell: Format | None = None,
        ) -> tp.Any:
            # cannot yet write complex types directly, so covert to string
            if isinstance(value, COMPLEX_TYPES):
                return ws.write_string(row, col, str(value), format_cell)

            if writer_attr == 'write':
                # determine type for each value
                if isinstance(value, BOOL_TYPES):
                    return ws.write_boolean(row, col, value, format_cell)  # pyright: ignore
                if isinstance(value, str):
                    return ws.write_string(row, col, value, format_cell)
                if isinstance(value, NUMERIC_TYPES):
                    return ws.write_number(row, col, value, format_cell)  # pyright: ignore
                if isinstance(
                    value, datetime.datetime
                ):  # NOTE: must come before date isinstance check
                    return ws.write_datetime(row, col, value, format_datetime)
                if isinstance(value, datetime.date):
                    return ws.write_datetime(row, col, value, format_date)  # pyright: ignore
            # use the type specific writer_native
            return writer_native(row, col, value, format_cell)

        return writer

    @classmethod
    def _frame_to_worksheet(
        cls,
        frame: TFrameAny,
        ws: 'Worksheet',
        *,
        format_columns: 'Format',
        format_index: 'Format',
        format_date: 'Format',
        format_datetime: 'Format',
        format_columns_date: 'Format',
        format_columns_datetime: 'Format',
        format_index_date: 'Format',
        format_index_datetime: 'Format',
        config: StoreConfigXLSX,
    ) -> None:
        c = config
        if sum((c.include_columns_name, c.include_index_name)) > 1:
            raise RuntimeError(
                'cannot set both `include_columns_name` and `include_index_name`'
            )

        index_depth = frame._index.depth
        index_depth_effective = 0 if not c.include_index else index_depth
        index_names = frame._index.names  # normalized presentation

        columns_iter = cls.get_column_iterator(frame=frame, include_index=c.include_index)

        columns_depth = frame._columns.depth
        columns_names = frame._columns.names
        columns_depth_effective = 0 if not c.include_columns else columns_depth

        columns_total = frame.shape[1] + index_depth_effective
        rows_total = frame.shape[0] + columns_depth_effective

        if rows_total > MAX_XLSX_ROWS:
            raise RuntimeError(
                f'Frame rows do not fit into XLSX sheet ({rows_total} > {MAX_XLSX_ROWS})'
            )
        if columns_total > MAX_XLSX_COLUMNS:
            raise RuntimeError(
                f'Frame columns do not fit into XLSX sheet ({columns_total} > {MAX_XLSX_COLUMNS})'
            )

        if c.include_columns:
            columns_values = frame._columns.values
            if c.store_filter:
                columns_values = c.store_filter.from_type_filter_array(columns_values)
            writer_columns = cls._get_writer(columns_values.dtype, ws)
            # for labels in apex, do not know type
            writer_names = cls._get_writer(DTYPE_OBJECT, ws)

        # write by column
        for col, values in enumerate(columns_iter):
            if c.include_columns:
                # The col integers will include index depth, so if including index, must wait until after index depth to write column field names; if include_index is False, can begin reading from columns_values
                if col < index_depth_effective:
                    if c.include_index_name:
                        writer_names(
                            0,  # always populate in top-most row
                            col,
                            index_names[col],
                            format_cell=format_index,
                            format_date=format_index_date,
                            format_datetime=format_index_datetime,
                        )
                    if c.include_columns_name and col == 0:
                        for i in range(columns_depth):
                            writer_names(
                                i,
                                col,  # always 0, populate in left-most colum
                                columns_names[i],
                                format_cell=format_columns,
                                format_date=format_columns_date,
                                format_datetime=format_columns_datetime,
                            )
                else:  # col >= index_depth_effective:
                    if columns_depth == 1:
                        writer_columns(
                            0,
                            col,
                            columns_values[col - index_depth_effective],
                            format_cell=format_columns,
                            format_date=format_columns_date,
                            format_datetime=format_columns_datetime,
                        )
                    elif columns_depth > 1:
                        for i in range(columns_depth):
                            # here, row selection is column count, column selection is depth
                            writer_columns(
                                i,
                                col,
                                columns_values[col - index_depth_effective, i],
                                format_cell=format_columns,
                                format_date=format_columns_date,
                                format_datetime=format_columns_datetime,
                            )
            if c.store_filter:
                # thi might change the dtype
                values = c.store_filter.from_type_filter_array(values)

            writer = cls._get_writer(values.dtype, ws)
            # start enumeration of row after the effective column depth
            for row, v in enumerate(values, columns_depth_effective):
                if col < index_depth_effective:
                    writer(
                        row,
                        col,
                        v,
                        format_cell=format_index,
                        format_date=format_index_date,
                        format_datetime=format_index_datetime,
                    )
                else:
                    writer(
                        row,
                        col,
                        v,
                        format_cell=None,
                        format_date=format_date,
                        format_datetime=format_datetime,
                    )

        # post process to merge cells; need to get width of at depth
        if c.include_columns and c.merge_hierarchical_labels and columns_depth > 1:
            for depth in range(columns_depth - 1):  # never most deep
                row = depth
                col = index_depth_effective  # start after index
                for label, width in frame._columns.label_widths_at_depth(depth):
                    # TODO: use store_filter
                    if width > 1:
                        ws.merge_range(
                            row, col, row, col + width - 1, label, format_columns
                        )
                    col += width

        if c.include_index and c.merge_hierarchical_labels and index_depth > 1:
            for depth in range(index_depth - 1):  # never most deep
                row = columns_depth_effective
                col = depth
                for label, width in frame._index.label_widths_at_depth(depth):
                    # TODO: use store_filter
                    if width > 1:
                        ws.merge_range(
                            row, col, row + width - 1, col, label, format_index
                        )
                    row += width

    @store_coherent_write
    def write(
        self,
        items: tp.Iterable[tuple[TLabel, TFrameAny]],
    ) -> None:
        # format_data: tp.Optional[tp.Dict[TLabel, tp.Dict[str, tp.Any]]]
        # format_data: dictionary of dictionaries, keyed by column label, that contains dictionaries of XlsxWriter format specifications.
        import xlsxwriter

        # NOTE: can supply second argument: {'default_date_format': 'dd/mm/yy'}
        with xlsxwriter.Workbook(self._fp, {'remove_timezone': True}) as wb:
            for label, frame in items:
                c = self._config[label]
                if label is STORE_LABEL_DEFAULT:
                    # None is supported by add_worksheet, below
                    label = None
                else:
                    label = self._config.default.label_encode(label)

                # NOTE: this must be called here, as we need the workbook been assigning formats, and we need to get a config per label
                format_columns = FormatDefaults.get_format_or_default(
                    wb, format_funcs=(FormatDefaults.label,)
                )
                format_index = FormatDefaults.get_format_or_default(
                    wb, format_funcs=(FormatDefaults.label,)
                )

                format_date = FormatDefaults.get_format_or_default(
                    wb, format_funcs=(FormatDefaults.date,)
                )
                format_datetime = FormatDefaults.get_format_or_default(
                    wb, format_funcs=(FormatDefaults.datetime,)
                )

                format_columns_date = FormatDefaults.get_format_or_default(
                    wb, format_funcs=(FormatDefaults.label, FormatDefaults.date)
                )
                format_columns_datetime = FormatDefaults.get_format_or_default(
                    wb,
                    format_funcs=(
                        FormatDefaults.label,
                        FormatDefaults.datetime,
                    ),
                )

                format_index_date = FormatDefaults.get_format_or_default(
                    wb, format_funcs=(FormatDefaults.label, FormatDefaults.date)
                )
                format_index_datetime = FormatDefaults.get_format_or_default(
                    wb,
                    format_funcs=(
                        FormatDefaults.label,
                        FormatDefaults.datetime,
                    ),
                )

                ws = wb.add_worksheet(label)  # label can be None
                self._frame_to_worksheet(
                    frame,
                    ws,
                    format_columns=format_columns,
                    format_index=format_index,
                    format_date=format_date,
                    format_datetime=format_datetime,
                    format_columns_date=format_columns_date,
                    format_columns_datetime=format_columns_datetime,
                    format_index_date=format_index_date,
                    format_index_datetime=format_index_datetime,
                    config=c,
                )

    @staticmethod
    def _load_workbook(fp: str) -> WorkbookOpenpyxl:
        import openpyxl

        # NOTE: read_only=True provides best performance, but may lead to empty cells with formatting being loaded
        return openpyxl.load_workbook(filename=fp, read_only=True, data_only=True)

    @store_coherent_non_write
    def read_many(
        self,
        labels: tp.Iterable[TLabel],
    ) -> tp.Iterator[TFrameAny]:
        wb = self._load_workbook(self._fp)

        for label in labels:
            c = self._config[label]

            if label is STORE_LABEL_DEFAULT:
                ws = wb[wb.sheetnames[0]]  # pyright: ignore
                name = None  # do not set to default sheet name
            else:
                label_encoded = self._config.default.label_encode(label)
                ws = wb[label_encoded]  # pyright: ignore
                name = label  # set name to the un-encoded hashable

            if ws.max_column <= 1 or ws.max_row <= 1:  # pyright: ignore
                # https://openpyxl.readthedocs.io/en/stable/optimized.html
                # says that some clients might not report correct dimensions
                ws.calculate_dimension()  # pyright: ignore

            max_column: int = ws.max_column  # pyright: ignore
            max_row: int = ws.max_row  # pyright: ignore

            # adjust for downward shift for skipping header, then reduce for footer; at this value and beyond we stop
            last_row_count: int = max_row - c.skip_header - c.skip_footer

            index_values = []
            columns_values: list[tp.Any] = []
            data = []
            apex_rows = []

            if c.trim_nadir:
                mask: TNDArray2DBool = np.full((last_row_count, max_column), False)

            for row_count, row in enumerate(
                ws.iter_rows(max_row=max_row),  # pyright: ignore
                start=-c.skip_header,
            ):
                if row_count < 0:
                    continue  # due to skip header; preserves comparison to columns_depth
                if row_count >= last_row_count:
                    break

                if c.trim_nadir:
                    row_data: tp.Sequence[tp.Any] = []
                    for col_count, cell in enumerate(row):
                        if c.store_filter is None:
                            value = cell.value
                        else:
                            value = c.store_filter.to_type_filter_element(cell.value)
                        if value is None:  # NOTE: only checking None, not np.nan
                            mask[row_count, col_count] = True
                        row_data.append(value)  # type: ignore
                    if not row_data:
                        # NOTE: there might be scenarios where there are empty ``row`` iterables that still increment the row_count; we cannot generate these directly for test
                        mask[row_count] = True  # pragma: no cover
                else:
                    if c.store_filter is None:
                        row_data = tuple(cell.value for cell in row)
                    else:  # only need to filter string values, but probably too expensive to pre-check
                        row_data = tuple(
                            c.store_filter.to_type_filter_element(cell.value)
                            for cell in row
                        )

                if row_count <= c.columns_depth - 1:
                    apex_rows.append(row_data[: c.index_depth])
                    if c.columns_depth == 1:
                        columns_values.extend(row_data[c.index_depth :])
                    elif c.columns_depth > 1:
                        columns_values.append(row_data[c.index_depth :])
                    continue

                if c.index_depth == 0:
                    data.append(row_data)
                elif c.index_depth == 1:
                    index_values.append(row_data[0])
                    data.append(row_data[1:])
                else:
                    index_values.append(row_data[: c.index_depth])
                    data.append(row_data[c.index_depth :])

            # -----------------------------------------------------------------------
            # Trim all-empty trailing rows created from style formatting GH#146. As the wb is opened in read-only mode, reverse iterating on the wb is not an option, nor is direct row access by integer
            if c.trim_nadir:
                # NOTE: `mask` is all data, while `data` is post index/columns extraction; this means that if a non-None label is found, the row/column will not be trimmed.
                row_mask: TNDArray1DBool = mask.all(axis=1)  # type: ignore
                row_trim_start = (
                    array1d_to_last_contiguous_to_edge(row_mask) - c.columns_depth
                )
                if row_trim_start < len(row_mask) - c.columns_depth:
                    data = data[:row_trim_start]
                    if c.index_depth > 0:  # this handles depth 1 and greater
                        index_values = index_values[:row_trim_start]

                col_mask: TNDArray1DBool = mask.all(axis=0)  # type: ignore
                col_trim_start = (
                    array1d_to_last_contiguous_to_edge(col_mask) - c.index_depth
                )
                if col_trim_start < len(col_mask) - c.index_depth:
                    data = (r[:col_trim_start] for r in data)  # type: ignore
                    if c.columns_depth == 1:
                        columns_values = columns_values[:col_trim_start]
                    if c.columns_depth > 1:
                        columns_values = (r[:col_trim_start] for r in columns_values)  # type: ignore

            # -----------------------------------------------------------------------
            # continue with Index and Frame creation
            index_name = (
                None
                if c.columns_depth == 0
                else apex_to_name(
                    rows=apex_rows,
                    depth_level=c.index_name_depth_level,
                    axis=0,
                    axis_depth=c.index_depth,
                )
            )

            # index: tp.Optional[IndexBase] = None
            index_default_constructor: TIndexCtor

            if c.index_depth <= 1:
                index_default_constructor = partial(Index, name=index_name)
            else:  # > 1
                index_default_constructor = partial(
                    IndexHierarchy.from_labels,
                    name=index_name,
                    continuation_token=None,  # NOTE: needed
                )

            index, own_index = index_from_optional_constructors(
                index_values,
                depth=c.index_depth,
                default_constructor=index_default_constructor,
                explicit_constructors=c.index_constructors,  # cannot supply name
            )

            columns_name = (
                None
                if c.index_depth == 0
                else apex_to_name(
                    rows=apex_rows,
                    depth_level=c.columns_name_depth_level,
                    axis=1,
                    axis_depth=c.columns_depth,
                )
            )

            # columns: tp.Optional[IndexBase] = None
            # own_columns = False
            columns_default_constructor: TIndexCtor
            if c.columns_depth <= 1:
                columns_default_constructor = partial(
                    Frame._COLUMNS_CONSTRUCTOR,
                    name=columns_name,
                )
            elif c.columns_depth > 1:
                columns_default_constructor = partial(
                    Frame._COLUMNS_HIERARCHY_CONSTRUCTOR.from_labels,
                    name=columns_name,
                    continuation_token=None,  # NOTE: needed, not the default
                )
                columns_values = zip(*columns_values)  # type: ignore

            columns, own_columns = index_from_optional_constructors(
                columns_values,
                depth=c.columns_depth,
                default_constructor=columns_default_constructor,
                explicit_constructors=c.columns_constructors,  # cannot supply name
            )

            f = Frame.from_records(
                data,
                index=index,
                columns=columns,
                dtypes=c.dtypes,
                own_index=own_index,
                own_columns=own_columns,
                name=name,
                consolidate_blocks=c.consolidate_blocks,
            )
            if c.read_frame_filter is not None:
                yield c.read_frame_filter(label, f)
            else:
                yield f

        wb.close()

    @store_coherent_non_write
    def labels(
        self,
        *,
        strip_ext: bool = True,
    ) -> tp.Iterator[TLabel]:
        wb = self._load_workbook(self._fp)
        labels = tuple(wb.sheetnames)
        wb.close()

        for label in labels:
            yield self._config.default.label_decode(label)
